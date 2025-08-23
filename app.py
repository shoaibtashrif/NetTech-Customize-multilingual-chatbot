import os
import json
import uuid
import logging
import requests
from datetime import datetime
from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from dotenv import load_dotenv
import codecs
from openai import OpenAI
from zoneinfo import ZoneInfo
import re
import pprint
import numpy as np
from sentence_transformers import SentenceTransformer
import faiss
import pickle
from typing import List, Dict, Any
import time
import random

# ─── Setup ─────────────────────────────────────────────────────────────
load_dotenv(override=True)

# OpenAI client
openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Groq API key - loaded from .env file
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
if GROQ_API_KEY:
    logging.info("GROQ_API_KEY loaded successfully from .env file")
else:
    logging.warning("GROQ_API_KEY not found in .env file - Groq API will not work")

# Admin‐controlled toggle
current_model = "openai"

app = Flask(__name__, template_folder="templates", static_folder="static")
CORS(app)

# Configure logging to write to both console and file
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("logs/app.log"),
        logging.StreamHandler()
    ]
)

# ─── Global Overrides ──────────────────────────────────────────────────
global_prompt    = None
global_knowledge = None

# ─── In‐Memory Stores ──────────────────────────────────────────────────
histories   = {}   # sid → list of {role,content}
last_active = {}   # sid → datetime
session_types = {} # sid → 'info' only (removed complaint)
session_type_history = {} # sid → set of types used in session

RECENT_LOGS = []
MAX_LOGS = 100

# ─── Token Logging ─────────────────────────────────────────────────────
def log_token_usage(model_name, input_tokens, output_tokens, total_cost=None):
    """Log token usage for each API call"""
    timestamp = now_str()
    
    # Calculate estimated cost if not provided
    if not total_cost:
        if "gpt-4" in model_name.lower():
            # GPT-4 pricing: $0.03 per 1K input tokens, $0.06 per 1K output tokens
            input_cost = (input_tokens / 1000) * 0.03
            output_cost = (output_tokens / 1000) * 0.06
            total_cost = input_cost + output_cost
        elif "groq" in model_name.lower() or "llama" in model_name.lower():
            # Groq pricing: $0.05 per 1M input tokens, $0.10 per 1M output tokens
            input_cost = (input_tokens / 1000000) * 0.05
            output_cost = (output_tokens / 1000000) * 0.10
            total_cost = input_cost + output_cost
        else:
            total_cost = 0.0
    
    log_entry = f"TOKEN_USAGE [{timestamp}] {model_name}: input={input_tokens}, output={output_tokens}, cost=${total_cost:.6f}"
    
    # Log to UI (for real-time display)
    log_to_ui(log_entry, 'info')
    
    # Log to file with detailed information
    detailed_log = f"TOKEN_USAGE_DETAILED [{timestamp}] Model: {model_name}, Input Tokens: {input_tokens}, Output Tokens: {output_tokens}, Cost: ${total_cost:.6f}"
    
    # Write to app.log file
    logging.info(detailed_log)
    
    # Also log to console for debugging
    print(f"🔢 {detailed_log}")

# ─── FAISS RAG Setup ──────────────────────────────────────────────────
# FAISS index and knowledge chunks are persisted to disk (faiss_index.pkl, knowledge_chunks.pkl)
# This means the RAG system maintains its knowledge base across app restarts
faiss_index = None
knowledge_chunks = []  # List of chunk dictionaries: {text, file_id, file_name, file_type, chunk_index}
file_metadata = {}     # file_id -> {name, type, upload_date, size}
embedding_model = None

def initialize_faiss():
    global faiss_index, embedding_model, knowledge_chunks, file_metadata
    try:
        # Load pre-trained sentence transformer model
        embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
        
        # Try to load existing FAISS index
        if os.path.exists('faiss_index.pkl') and os.path.exists('knowledge_chunks.pkl'):
            with open('faiss_index.pkl', 'rb') as f:
                faiss_index = pickle.load(f)
            with open('knowledge_chunks.pkl', 'rb') as f:
                loaded_chunks = pickle.load(f)
            
            # Handle backward compatibility
            if loaded_chunks and isinstance(loaded_chunks[0], dict):
                # New format with metadata
                knowledge_chunks = loaded_chunks
                # Load file metadata if exists
                if os.path.exists('file_metadata.pkl'):
                    with open('file_metadata.pkl', 'rb') as f:
                        file_metadata = pickle.load(f)
                else:
                    file_metadata = {}
            else:
                # Old format - convert to new format
                knowledge_chunks = [{"text": chunk, "file_id": "legacy", "file_name": "custom_knowledge.txt", "file_type": "txt", "chunk_index": i} for i, chunk in enumerate(loaded_chunks)]
                file_metadata = {"legacy": {"name": "custom_knowledge.txt", "type": "txt", "upload_date": "2024-01-01", "size": 0}}
            
            logging.info(f"Loaded existing FAISS index with {len(knowledge_chunks)} chunks from {len(file_metadata)} files")
        else:
            # Create new index
            dimension = embedding_model.get_sentence_embedding_dimension()
            faiss_index = faiss.IndexFlatIP(dimension)
            knowledge_chunks = []
            file_metadata = {}
            logging.info("Created new FAISS index")
    except Exception as e:
        logging.error(f"FAISS initialization failed: {e}")
        # Fallback to empty state
        faiss_index = None
        knowledge_chunks = []
        file_metadata = {}

def chunk_knowledge(text: str, file_id: str, file_name: str, file_type: str, chunk_size: int = 500, overlap: int = 50) -> List[Dict[str, Any]]:
    """Split knowledge base into overlapping chunks with metadata"""
    if not text.strip():
        return []
    
    chunks = []
    sentences = re.split(r'[.!?]+', text)
    current_chunk = ""
    chunk_index = 0
    
    for sentence in sentences:
        sentence = sentence.strip()
        if not sentence:
            continue
            
        if len(current_chunk) + len(sentence) < chunk_size:
            current_chunk += sentence + ". "
        else:
            if current_chunk:
                chunks.append({
                    "text": current_chunk.strip(),
                    "file_id": file_id,
                    "file_name": file_name,
                    "file_type": file_type,
                    "chunk_index": chunk_index
                })
                chunk_index += 1
            current_chunk = sentence + ". "
    
    if current_chunk:
        chunks.append({
            "text": current_chunk.strip(),
            "file_id": file_id,
            "file_name": file_name,
            "file_type": file_type,
            "chunk_index": chunk_index
        })
    
    return chunks

def update_faiss_index(knowledge_text: str, file_id: str = None, file_name: str = None, file_type: str = None, replace_all: bool = False):
    """Update FAISS index with new knowledge base"""
    global faiss_index, knowledge_chunks, file_metadata
    
    if not embedding_model:
        initialize_faiss()
    
    # If replace_all is True, clear everything and start fresh
    if replace_all:
        knowledge_chunks = []
        file_metadata = {}
        if faiss_index:
            faiss_index.reset()
    
    # Create chunks with metadata
    if file_id and file_name and file_type:
        new_chunks = chunk_knowledge(knowledge_text, file_id, file_name, file_type)
        # Add file metadata
        file_metadata[file_id] = {
            "name": file_name,
            "type": file_type,
            "upload_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "size": len(knowledge_text)
        }
    else:
        # Legacy support - create chunks with default metadata
        file_id = "legacy"
        file_name = "custom_knowledge.txt"
        file_type = "txt"
        new_chunks = chunk_knowledge(knowledge_text, file_id, file_name, file_type)
        file_metadata[file_id] = {
            "name": file_name,
            "type": file_type,
            "upload_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "size": len(knowledge_text)
        }
    
    if not new_chunks:
        return
    
    # Add new chunks to existing ones
    knowledge_chunks.extend(new_chunks)
    
    # Create embeddings for all chunks
    chunk_texts = [chunk["text"] for chunk in knowledge_chunks]
    embeddings = embedding_model.encode(chunk_texts)
    
    # Update index
    if faiss_index is None:
        dimension = embedding_model.get_sentence_embedding_dimension()
        faiss_index = faiss.IndexFlatIP(dimension)
    
    # Clear existing data and add all chunks
    faiss_index.reset()
    faiss_index.add(embeddings.astype('float32'))
    
    # Save to disk
    try:
        with open('faiss_index.pkl', 'wb') as f:
            pickle.dump(faiss_index, f)
        with open('knowledge_chunks.pkl', 'wb') as f:
            pickle.dump(knowledge_chunks, f)
        with open('file_metadata.pkl', 'wb') as f:
            pickle.dump(file_metadata, f)
        logging.info(f"Updated FAISS index with {len(knowledge_chunks)} total chunks from {len(file_metadata)} files")
    except Exception as e:
        logging.error(f"Failed to save FAISS index: {e}")

def search_knowledge(query: str, top_k: int = 3) -> List[Dict[str, Any]]:
    """Search knowledge base using FAISS and return chunks with metadata"""
    global faiss_index, knowledge_chunks, embedding_model
    
    if not faiss_index or not knowledge_chunks or not embedding_model:
        return []
    
    try:
        # Create query embedding
        query_embedding = embedding_model.encode([query])
        
        # Search
        scores, indices = faiss_index.search(query_embedding.astype('float32'), top_k)
        
        # Return relevant chunks with metadata
        results = []
        for i, idx in enumerate(indices[0]):
            if idx < len(knowledge_chunks):
                chunk = knowledge_chunks[idx].copy()
                chunk["similarity_score"] = float(scores[0][i])
                results.append(chunk)
        
        return results
    except Exception as e:
        logging.error(f"FAISS search failed: {e}")
        return []

def delete_file_from_knowledge(file_id: str) -> bool:
    """Delete a file from the knowledge base and rebuild FAISS index"""
    global faiss_index, knowledge_chunks, file_metadata
    
    if file_id not in file_metadata:
        logging.warning(f"File ID {file_id} not found in knowledge base")
        return False
    
    try:
        # Remove chunks from this file
        knowledge_chunks = [chunk for chunk in knowledge_chunks if chunk["file_id"] != file_id]
        
        # Remove file metadata
        deleted_file = file_metadata.pop(file_id)
        
        # Rebuild FAISS index with remaining chunks
        if knowledge_chunks:
            chunk_texts = [chunk["text"] for chunk in knowledge_chunks]
            embeddings = embedding_model.encode(chunk_texts)
            
            # Create new index
            dimension = embedding_model.get_sentence_embedding_dimension()
            faiss_index = faiss.IndexFlatIP(dimension)
            faiss_index.add(embeddings.astype('float32'))
            
            # Save updated index
            with open('faiss_index.pkl', 'wb') as f:
                pickle.dump(faiss_index, f)
            with open('knowledge_chunks.pkl', 'wb') as f:
                pickle.dump(knowledge_chunks, f)
            with open('file_metadata.pkl', 'wb') as f:
                pickle.dump(file_metadata, f)
            
            logging.info(f"Deleted file '{deleted_file['name']}' and rebuilt index with {len(knowledge_chunks)} chunks from {len(file_metadata)} files")
        else:
            # No chunks left, reset everything
            faiss_index = None
            with open('faiss_index.pkl', 'wb') as f:
                pickle.dump(None, f)
            with open('knowledge_chunks.pkl', 'wb') as f:
                pickle.dump([], f)
            with open('file_metadata.pkl', 'wb') as f:
                pickle.dump({}, f)
            logging.info("Deleted all files, reset knowledge base")
        
        return True
        
    except Exception as e:
        logging.error(f"Failed to delete file {file_id}: {e}")
        return False

def get_knowledge_base_files() -> List[Dict[str, Any]]:
    """Get list of all files in the knowledge base with metadata"""
    global file_metadata
    
    files = []
    for file_id, metadata in file_metadata.items():
        # Count chunks for this file
        chunk_count = len([chunk for chunk in knowledge_chunks if chunk["file_id"] == file_id])
        
        file_info = {
            "file_id": file_id,
            "name": metadata["name"],
            "type": metadata["type"],
            "upload_date": metadata["upload_date"],
            "size": metadata["size"],
            "chunk_count": chunk_count
        }
        files.append(file_info)
    
    return files

# ─── Program Q&A Knowledge Base ───────────────────────────────────────
PROGRAM_QA = {
    # Eligibility questions
    "age requirement": {
        "answer": "Age for this program is **13-19 years** girls",
        "variations": ["what is the age limit", "how old should the girl be", "age criteria"]
    },
    "registration criteria": {
        "answer": "Requirements: **Mother CNIC**, girl's **B-form**, and age **13-19 years**",
        "variations": ["what documents needed", "registration requirements", "what is needed to enroll"]
    },
    
    # Process questions
    "first visit process": {
        "answer": "First visit steps: 1) **Register**, 2) **Health check**, 3) **Awareness session**, 4) Receive **IFA**",
        "variations": ["what happens in first visit", "initial visit process", "first time procedure"]
    },
    "cash transfer amount": {
        "answer": "Cash transfer: **Rs. 2000/- per quarter** when compliance met",
        "variations": ["how much money", "payment amount", "cash benefit"]
    },
    
    # Special cases
    "disabled girl registration": {
        "answer": "Disabled girls: Can register if they have **B-form** and visit FC with mother",
        "variations": ["what if disabled", "handicapped girl", "special needs registration"]
    },
    "multiple daughters": {
        "answer": "Multiple daughters: All eligible if they have **B-forms** (no limit per family)",
        "variations": ["two daughters", "sisters enrollment", "more than one girl"]
    },
    "mother not present": {
        "answer": "**Mother must be present** - required for cash transfer and compliance",
        "variations": ["father brings girl", "mother not available", "can father come instead"]
    },
    "district requirement": {
        "answer": "Only these districts: **Neelam**, **Lasbela**, **Hub**, **Qambar Shahdadkot**, **Swat**, **Rajanpur**, **Jampur**, **Ghizer**",
        "variations": ["which areas eligible", "my district included", "location requirements"]
    },
    "ifa supplementation": {
        "answer": "IFA protocol: **60mg iron + 2800μg folic acid**, once weekly for **3 months**",
        "variations": ["what tablets given", "supplement details", "medicine protocol"]
    },
    "missed dose": {
        "answer": "Missed dose: Take next day if remembered, otherwise wait for next week's dose",
        "variations": ["forgot tablet", "missed medicine", "what if not taken"]
    },
    "side effects": {
        "answer": "Side effects: **Consult doctor** if severe issues, take with water after meals",
        "variations": ["nausea from tablets", "constipation issues", "medicine problems"]
    },
    
    # New Critical FAQs
    "adolescent nutrition program": {
        "answer": "The Adolescent Nutrition Program (ANP) focuses on improving nutritional status of **adolescent girls aged 13-19 years** in targeted districts. It aims to break the intergenerational cycle of malnutrition, reduce early pregnancy risks, and ensure healthier future mothers.",
        "variations": ["what is anp", "adolescent nutrition component", "what is the program about"]
    },
    "who is eligible": {
        "answer": "**Adolescent girls aged 13 years to 17 years, 11 months, and 29 days** who are registered as BISP beneficiaries or part of the **8171 program** (above PMT score). At registration, the girl must be **unmarried**, but can continue if married later.",
        "variations": ["eligibility criteria", "who can join", "who can register", "who qualifies"]
    },
    "what support provided": {
        "answer": "Services include: **IFA tablets** to prevent anemia, **balanced diet education**, **menstrual hygiene management**, **physical activity guidance**, **reproductive health awareness**, **counseling for healthy behaviors**, **peer group sessions**, and **community awareness activities**.",
        "variations": ["what services", "what help provided", "what support given", "what do they provide"]
    },
    "how to enroll": {
        "answer": "Visit the nearest **BNP Facilitation Center** (often in RHCs, THQs, or DHQs) in **Rajanpur**. Staff will record: **basic personal details** and **health information** (weight, height, MUAC). Records are maintained manually or via the **BISP Nashonuma App**.",
        "variations": ["how to register", "enrollment process", "how to join", "registration process"]
    },
    "is there cost": {
        "answer": "**No cost** - There's no fee for registration or services under BNP.",
        "variations": ["is it free", "how much does it cost", "registration fee", "service charges"]
    },
    "why adolescent nutrition important": {
        "answer": "Adolescents (**13-19 years**) go through fast physical growth, hormonal changes, and brain development requiring extra nutrients. Many girls suffer from **anemia and micronutrient deficiencies**. Good nutrition prevents these problems and supports healthy development.",
        "variations": ["why is it important", "why nutrition matters", "why adolescent nutrition", "importance of nutrition"]
    },
    "school going girls": {
        "answer": "The program is for **all eligible adolescent girls (13-19 years)**, whether they are **in school or out of school**.",
        "variations": ["can school girls join", "students eligible", "school attendance required", "education requirement"]
    },
    "visit frequency": {
        "answer": "Adolescent girls are expected to visit the center **every 3 months (quarterly)**.",
        "variations": ["how often to visit", "visit schedule", "frequency of visits", "how many times to come"]
    },
    "are boys included": {
        "answer": "**No, boys are not included**. Currently, the adolescent nutrition component focuses mainly on **adolescent girls**.",
        "variations": ["can boys join", "male eligibility", "boys allowed", "gender requirement"]
    },
    "documents required": {
        "answer": "Required documents: **B-Form of adolescent** updated with **Mother's CNIC** and **Mother's CNIC**.",
        "variations": ["what documents needed", "required papers", "documentation needed", "papers required"]
    },
    "married adolescent": {
        "answer": "At registration, the adolescent girl must be **unmarried**. However, if she gets married **after enrollment** and is willing to continue, she remains eligible to receive services.",
        "variations": ["what if married", "married girl eligibility", "marriage status", "can married girls join"]
    },
    "multiple sisters": {
        "answer": "**Yes, up to four sisters** can be enrolled in the program simultaneously.",
        "variations": ["sisters enrollment", "multiple daughters", "family limit", "how many sisters"]
    },
    "mother passed away": {
        "answer": "If an adolescent's mother has passed away, the adolescent can choose whether to continue. If yes, then we will give **only IFAS**.",
        "variations": ["mother died", "orphaned girl", "no mother", "mother not available"]
    },
    "district migration": {
        "answer": "This is currently a **pilot project limited to Rajanpur district**. If BISP beneficiaries migrate to another district, they should visit the local BISP office to update their district of residence for continuity of services.",
        "variations": ["moving to another district", "district change", "relocation", "different district"]
    },
    "facilitation center facilities": {
        "answer": "Facilitation Centres provide: **IEC materials**, **disposable glasses**, **session calendar**, **social maps**, **CFM numbers**, and **emergency contact numbers**.",
        "variations": ["what facilities available", "center facilities", "what's at the center", "center amenities"]
    },
    "system issues": {
        "answer": "Known system issues: Some graduated girls still appear in due list, exited beneficiaries remain visible, application shows errors (referred to IT admin), premature quarter skipping allowed, and above PMT beneficiaries account creation pending.",
        "variations": ["app problems", "system errors", "technical issues", "application problems"]
    }
}

# ─── Helpers ───────────────────────────────────────────────────────────
def now_str():
    return datetime.now(ZoneInfo("Asia/Karachi")).strftime("%Y-%m-%d %H:%M:%S")

def extract_text(f):
    return f.read().decode("utf-8", errors="ignore")

def load_prompt(sid=None):
    if global_prompt:
        return global_prompt
    if os.path.exists("system_prompt.txt"):
        return codecs.open("system_prompt.txt", "r", "utf8").read()
    return "Default system prompt."

def load_knowledge(sid=None):
    if global_knowledge:
        return global_knowledge
    if os.path.exists("custom_knowledge.txt"):
        try:
            return codecs.open("custom_knowledge.txt", "r", "utf8").read()
        except Exception as e:
            logging.error(f"Error loading knowledge base: {str(e)}")
            return ""
    return "Knowledge base not found. Please upload knowledge base."

def ai_summarize(chat):
    if current_model == "openai":
        try:
            resp = openai_client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role":"system","content":"Summarize this conversation in English."},
                    {"role":"user","content":"\n".join(f"{m['role']}: {m['content']}" for m in chat)}
                ]
            )
            # Log token usage
            input_tokens = resp.usage.prompt_tokens
            output_tokens = resp.usage.completion_tokens
            log_token_usage("OpenAI GPT-4", input_tokens, output_tokens)
            return resp.choices[0].message.content
        except Exception as e:
            logging.error(f"OpenAI summarization failed: {e}")
            return "Conversation summary unavailable."
    elif current_model == "groq":
        groq_url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {GROQ_API_KEY}"
        }
        payload = {
            "model": "llama3-8b-8192",
            "messages": [
                {"role":"system","content":"Summarize this conversation in English."},
                {"role":"user","content":"\n".join(f"{m['role']}: {m['content']}" for m in chat)}
            ]
        }
        
        # Use retry logic for Groq API calls
        resp = make_groq_request_with_retry(groq_url, headers, payload)
        
        if resp is None:
            logging.error("Groq summarization failed after all retries")
            return "Conversation summary unavailable."
        elif resp.status_code != 200:
            logging.error(f"Groq summarization failed: {resp.status_code} - {resp.text}")
            return "Conversation summary unavailable."
        else:
            try:
                data = resp.json()
                # Log token usage for Groq
                input_tokens = data.get("usage", {}).get("prompt_tokens", 0)
                output_tokens = data.get("usage", {}).get("completion_tokens", 0)
                log_token_usage("Groq Llama3-8B", input_tokens, output_tokens)
                return data["choices"][0]["message"]["content"]
            except Exception as e:
                logging.error(f"Groq summarization parsing failed: {e}")
                return "Conversation summary unavailable."
    else:
        # Fallback - simple concatenation
        return " ".join([m['content'] for m in chat if m['role'] == 'user'])[:200] + "..."

def log_to_ui(msg, level='info'):
    global RECENT_LOGS
    RECENT_LOGS.append(msg)
    if len(RECENT_LOGS) > MAX_LOGS:
        RECENT_LOGS = RECENT_LOGS[-MAX_LOGS:]
    if level == 'info':
        logging.info(msg)
    elif level == 'warning':
        logging.warning(msg)
    elif level == 'error':
        logging.error(msg)
    else:
        logging.info(msg)

def ai_detect_types(chat, session_types_map=None):
    """Simplified function that only handles info sessions - no complaints"""
    log_to_ui(f"AI_DETECT_TYPES - Starting analysis of {len(chat)} messages (INFO ONLY)")
    
    # All sessions are now info sessions
    info_categories = []
    
    # Process each user message to categorize it
    for idx, m in enumerate(chat):
        if m.get('role') == 'user':
            content = m.get('content', '')
            log_to_ui(f"AI_DETECT_TYPES - Processing message {idx}: '{content[:50]}...' (type: info)")
            
            # Simple keyword-based categorization for info sessions
            categories = categorize_info_message(content)
            info_categories.extend(categories)
    
    # Remove duplicates while preserving order
    def dedup_and_order(lst):
        seen = set()
        ordered = []
        if 'money' in lst:
            ordered.append('money')
            seen.add('money')
        for i in lst:
            if i not in seen and i in ['money', 'eligibility', 'district', 'ingredients', 'other']:
                ordered.append(i)
                seen.add(i)
        return ordered
    
    info = dedup_and_order(info_categories)
    
    # CRITICAL: Ensure we always have at least one category
    if not info:
        info = ['other']
        log_to_ui(f"AI_DETECT_TYPES - No categories detected, defaulting to: {info}")
    
    log_to_ui(f"AI_DETECT_TYPES - Final result - info: {info}")
    return {'complaints': [], 'info': info}

def categorize_info_message(content):
    """Categorize a single info message using AI instead of unreliable keywords"""
    try:
        # Create a simple categorization prompt for the LLM
        categorization_prompt = f"""
        Categorize this message into ONE category:
        
        - money: Cash, payments, benefits, financial assistance
        - eligibility: Registration, requirements, criteria, age limits
        - district: Locations, areas, addresses, where to go
        - ingredients: Medicine, supplements, IFA tablets, vitamins
        - other: Any other questions

        Message: "{content}"
        
        Return ONLY the category name.
        """

        if current_model == "groq":
            # Use Groq for categorization (cheaper)
            groq_url = "https://api.groq.com/openai/v1/chat/completions"
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {GROQ_API_KEY}"
            }
            payload = {
                "model": "llama3-8b-8192",
                "messages": [
                    {"role": "system", "content": "You are a categorization expert. Return only the category name."},
                    {"role": "user", "content": categorization_prompt}
                ],
                "temperature": 0.1,
                "max_tokens": 20
            }
            
            resp = make_groq_request_with_retry(groq_url, headers, payload)
            if resp and resp.status_code == 200:
                data = resp.json()
                category = data["choices"][0]["message"]["content"].strip().lower()
                
                # Log token usage for categorization
                input_tokens = data.get("usage", {}).get("prompt_tokens", 0)
                output_tokens = data.get("usage", {}).get("completion_tokens", 0)
                log_token_usage("Groq Llama3-8B (categorization)", input_tokens, output_tokens)
                
                # Validate category
                valid_categories = ['money', 'eligibility', 'district', 'ingredients', 'other']
                if category in valid_categories:
                    return [category]
                else:
                    logging.warning(f"AI returned invalid category: {category}, defaulting to 'other'")
                    return ['other']
            else:
                logging.error("Groq categorization failed, defaulting to 'other'")
                return ['other']
                
        else:
            # Use OpenAI for categorization
            try:
                response = openai_client.chat.completions.create(
                    model="gpt-4",
                    messages=[
                        {"role": "system", "content": "You are a categorization expert. Return only the category name."},
                        {"role": "user", "content": categorization_prompt}
                    ],
                    temperature=0.1,
                    max_tokens=20
                )
                
                category = response.choices[0].message.content.strip().lower()
                
                # Log token usage for categorization
                input_tokens = response.usage.prompt_tokens
                output_tokens = response.usage.completion_tokens
                log_token_usage("OpenAI GPT-4 (categorization)", input_tokens, output_tokens)
                
                # Validate category
                valid_categories = ['money', 'eligibility', 'district', 'ingredients', 'other']
                if category in valid_categories:
                    return [category]
                else:
                    logging.warning(f"AI returned invalid category: {category}, defaulting to 'other'")
                    return ['other']
                    
            except Exception as e:
                logging.error(f"OpenAI categorization failed: {e}, defaulting to 'other'")
                return ['other']
                
    except Exception as e:
        logging.error(f"AI categorization completely failed: {e}, defaulting to 'other'")
        return ['other']

def find_matching_question(user_input):
    user_input = user_input.lower().strip()
    
    # First check for exact matches (highest priority)
    for q, data in PROGRAM_QA.items():
        if q == user_input:
            return data["answer"]
    
    # Then check for exact phrase matches in variations
    for q, data in PROGRAM_QA.items():
        for variation in data.get("variations", []):
            if variation == user_input:
                return data["answer"]
    
    # Then check if the question contains key phrases (medium priority)
    for q, data in PROGRAM_QA.items():
        # Split question into key words and check if most are present
        question_words = q.split()
        if len(question_words) >= 2:  # Only for questions with 2+ words
            matching_words = sum(1 for word in question_words if word in user_input)
            if matching_words >= len(question_words) * 0.7:  # 70% of words must match
                return data["answer"]
    
    # Then check variations with key phrase matching
    for q, data in PROGRAM_QA.items():
        for variation in data.get("variations", []):
            variation_words = variation.split()
            if len(variation_words) >= 2:
                matching_words = sum(1 for word in variation_words if word in user_input)
                if matching_words >= len(variation_words) * 0.7:
                    return data["answer"]
    
    # Finally, check for any word overlap (lowest priority)
    for q, data in PROGRAM_QA.items():
        question_words = q.split()
        if any(word in user_input for word in question_words):
            # Additional check: make sure it's not just a common word like "girl", "program", etc.
            common_words = ['girl', 'program', 'what', 'how', 'can', 'is', 'are', 'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by']
            meaningful_matches = [word for word in question_words if word not in common_words and word in user_input]
            if len(meaningful_matches) >= 2:  # At least 2 meaningful words must match
                return data["answer"]
    
    return None

def make_groq_request_with_retry(url, headers, payload, max_retries=3, base_delay=1):
    """Make a Groq API request with exponential backoff retry logic"""
    for attempt in range(max_retries):
        try:
            # Add jitter to avoid thundering herd
            delay = base_delay * (2 ** attempt) + random.uniform(0, 1)
            
            resp = requests.post(url, headers=headers, json=payload, timeout=45)  # Increased timeout
            
            if resp.status_code == 200:
                return resp
            elif resp.status_code == 429:  # Rate limit
                logging.warning(f"Groq rate limit hit, attempt {attempt + 1}/{max_retries}")
                if attempt < max_retries - 1:
                    time.sleep(delay)
                    continue
            else:
                logging.error(f"Groq API error: {resp.status_code} - {resp.text}")
                return resp
                
        except requests.exceptions.Timeout:
            logging.warning(f"Groq timeout on attempt {attempt + 1}/{max_retries}")
            if attempt < max_retries - 1:
                time.sleep(delay)
                continue
        except requests.exceptions.ConnectionError as e:
            logging.warning(f"Groq connection error on attempt {attempt + 1}/{max_retries}: {e}")
            if attempt < max_retries - 1:
                time.sleep(delay)
                continue
        except Exception as e:
            logging.error(f"Groq unexpected error on attempt {attempt + 1}/{max_retries}: {e}")
            if attempt < max_retries - 1:
                time.sleep(delay)
                continue
    
    # If all retries failed, return None
    logging.error(f"Groq API failed after {max_retries} attempts")
    return None

def extract_text_from_file(file_path: str, file_type: str) -> str:
    """Extract text from different file types"""
    try:
        if file_type.lower() in ['pdf']:
            return extract_text_from_pdf(file_path)
        elif file_type.lower() in ['doc', 'docx']:
            return extract_text_from_doc(file_path)
        elif file_type.lower() in ['xls', 'xlsx']:
            return extract_text_from_excel(file_path)
        elif file_type.lower() in ['txt']:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        else:
            logging.error(f"Unsupported file type: {file_type}")
            return ""
    except Exception as e:
        logging.error(f"Failed to extract text from {file_path}: {e}")
        return ""

def extract_text_from_uploaded_file(file, file_type: str) -> str:
    """Extract text from uploaded file object"""
    try:
        # Save uploaded file temporarily
        import tempfile
        import os
        
        # Create temp file with proper extension
        temp_fd, temp_path = tempfile.mkstemp(suffix=f".{file_type}")
        os.close(temp_fd)
        
        # Save uploaded file
        file.save(temp_path)
        
        # Extract text
        text = extract_text_from_file(temp_path, file_type)
        
        # Clean up temp file
        os.unlink(temp_path)
        
        return text
    except Exception as e:
        logging.error(f"Failed to extract text from uploaded file: {e}")
        return ""

def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF file"""
    try:
        import PyPDF2
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    except ImportError:
        logging.error("PyPDF2 not installed. Install with: pip install PyPDF2")
        return ""
    except Exception as e:
        logging.error(f"PDF extraction failed: {e}")
        return ""

def extract_text_from_doc(file_path: str) -> str:
    """Extract text from DOC/DOCX file"""
    try:
        from docx import Document
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except ImportError:
        logging.error("python-docx not installed. Install with: pip install python-docx")
        return ""
    except Exception as e:
        logging.error(f"DOC extraction failed: {e}")
        return ""

def extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel file"""
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        return text
    except ImportError:
        logging.error("openpyxl not installed. Install with: pip install openpyxl")
        return ""
    except Exception as e:
        logging.error(f"Excel extraction failed: {e}")
        return ""

# ─── Admin endpoints ───────────────────────────────────────────────────
@app.route("/upload", methods=["POST"])
def upload_override():
    global global_prompt, global_knowledge
    
    # Handle prompt file upload
    if "promptFile" in request.files:
        file = request.files["promptFile"]
        text = extract_text(file)
        global_prompt = text
        try:
            with open("system_prompt.txt", "w", encoding="utf8") as f:
                f.write(text)
            logging.info("Global prompt overridden and saved to system_prompt.txt.")
        except Exception as e:
            logging.error(f"Failed to save system_prompt.txt: {e}")
    
    # Handle knowledge file uploads
    if "knowledgeFile" in request.files:
        file = request.files["knowledgeFile"]
        file_id = str(uuid.uuid4())
        file_name = file.filename
        file_type = file_name.split('.')[-1].lower() if '.' in file_name else 'txt'
        
        # Extract text based on file type
        text = extract_text_from_uploaded_file(file, file_type)
        if not text:
            return jsonify({"status": "error", "message": f"Failed to extract text from {file_name}"}), 400
        
        # Update FAISS index with new file
        update_faiss_index(text, file_id, file_name, file_type)
        
        # Also update global knowledge for backward compatibility
        global_knowledge = text
        try:
            with open("custom_knowledge.txt", "w", encoding="utf8") as f:
                f.write(text)
            logging.info(f"Knowledge file '{file_name}' uploaded and FAISS index updated.")
        except Exception as e:
            logging.error(f"Failed to save custom_knowledge.txt: {e}")
    
    # Handle multiple knowledge files
    if "knowledgeFiles" in request.files:
        files = request.files.getlist("knowledgeFiles")
        uploaded_files = []
        
        for file in files:
            if file.filename:
                file_id = str(uuid.uuid4())
                file_name = file.filename
                file_type = file_name.split('.')[-1].lower() if '.' in file_name else 'txt'
                
                # Extract text based on file type
                text = extract_text_from_uploaded_file(file, file_type)
                if text:
                    # Update FAISS index with new file
                    update_faiss_index(text, file_id, file_name, file_type)
                    uploaded_files.append({
                        "file_id": file_id,
                        "name": file_name,
                        "type": file_type,
                        "status": "success"
                    })
                    logging.info(f"Knowledge file '{file_name}' uploaded and FAISS index updated.")
                else:
                    uploaded_files.append({
                        "name": file_name,
                        "type": file_type,
                        "status": "error",
                        "message": "Failed to extract text"
                    })
        
        return jsonify({
            "status": "files_uploaded",
            "uploaded_files": uploaded_files,
            "total_files": len(files)
        })
    
    return jsonify({"status": "override_uploaded"})

@app.route("/set_model", methods=["POST"])
def set_model():
    global current_model
    data = request.get_json() or {}
    m = data.get("model")
    if m not in ("openai", "groq"):
        return jsonify(status="error", message="Invalid model"), 400
    current_model = m
    logging.info(f"Global model switched to: {current_model}")
    return jsonify(status="ok", model=current_model)

@app.route("/files", methods=["GET"])
def list_files():
    """Get list of all files in the knowledge base"""
    try:
        files = get_knowledge_base_files()
        return jsonify({"status": "success", "files": files})
    except Exception as e:
        logging.error(f"Failed to list files: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/files/<file_id>", methods=["DELETE"])
def delete_file(file_id):
    """Delete a file from the knowledge base"""
    try:
        success = delete_file_from_knowledge(file_id)
        if success:
            return jsonify({"status": "success", "message": "File deleted successfully"})
        else:
            return jsonify({"status": "error", "message": "File not found or deletion failed"}), 404
    except Exception as e:
        logging.error(f"Failed to delete file {file_id}: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/files/clear", methods=["POST"])
def clear_all_files():
    """Clear all files from the knowledge base"""
    try:
        global faiss_index, knowledge_chunks, file_metadata
        
        # Reset everything
        faiss_index = None
        knowledge_chunks = []
        file_metadata = {}
        
        # Save empty state
        with open('faiss_index.pkl', 'wb') as f:
            pickle.dump(None, f)
        with open('knowledge_chunks.pkl', 'wb') as f:
            pickle.dump([], f)
        with open('file_metadata.pkl', 'wb') as f:
            pickle.dump({}, f)
        
        logging.info("All files cleared from knowledge base")
        return jsonify({"status": "success", "message": "All files cleared successfully"})
    except Exception as e:
        logging.error(f"Failed to clear all files: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/upload_prompt", methods=["POST"])
def upload_prompt():
    """Upload and update the system prompt file"""
    try:
        if 'prompt_file' not in request.files:
            return jsonify(status="error", message="No file uploaded"), 400
        
        file = request.files['prompt_file']
        if file.filename == '':
            return jsonify(status="error", message="No file selected"), 400
        
        if file and file.filename.endswith('.txt'):
            # Save the new prompt file
            file.save('system_prompt.txt')
            
            # Log the prompt update
            log_to_ui(f"SYSTEM_PROMPT_UPDATED - New prompt file uploaded: {file.filename}")
            logging.info(f"SYSTEM_PROMPT_UPDATED [{now_str()}] File: {file.filename}, Size: {len(file.read())} bytes")
            
            return jsonify(status="success", message="System prompt updated successfully")
        else:
            return jsonify(status="error", message="Only .txt files are allowed"), 400
            
    except Exception as e:
        log_to_ui(f"PROMPT_UPLOAD_ERROR - {str(e)}")
        logging.error(f"PROMPT_UPLOAD_ERROR [{now_str()}] {str(e)}")
        return jsonify(status="error", message=f"Error updating prompt: {str(e)}"), 500

@app.route("/get_prompt", methods=["GET"])
def get_prompt():
    """Get the current system prompt content"""
    try:
        with open('system_prompt.txt', 'r', encoding='utf-8') as f:
            prompt_content = f.read()
        
        return jsonify(status="success", prompt=prompt_content)
        
    except FileNotFoundError:
        return jsonify(status="error", message="System prompt file not found"), 404
    except Exception as e:
        log_to_ui(f"PROMPT_READ_ERROR - {str(e)}")
        logging.error(f"PROMPT_READ_ERROR [{now_str()}] {str(e)}")
        return jsonify(status="error", message=f"Error reading prompt: {str(e)}"), 500

# ─── UI routes ─────────────────────────────────────────────────────────
@app.route("/")
def home():
    return render_template("index.html")

@app.route("/client")
def client():
    return render_template("client.html")

# ─── Session start/end ─────────────────────────────────────────────────
@app.route("/start_toggle", methods=["POST"])
def start_toggle():
    data = request.get_json() or {}
    sid  = data.get("session_id")
    logging.info(f"Toggle session called with session_id={sid}")

    if not sid or sid not in histories:
        sid = str(uuid.uuid4())
        histories[sid] = []
        last_active[sid] = datetime.now()
        session_type_history[sid] = set()
        return jsonify({"started":True, "session_id":sid})

    chat     = histories.pop(sid, [])
    last_active.pop(sid, None)
    summary  = ai_summarize(chat)
    
    # Build session_types_map for ai_detect_types
    session_types_map = {}
    for idx, m in enumerate(chat):
        if m.get('role') == 'user':
            session_types_map[idx] = m.get('session_type', 'info')
    
    # Get final categorization for the entire session
    detected_types = ai_detect_types(chat, session_types_map)
    info = detected_types.get("info", [])
    
    # CRITICAL: Ensure every session has at least one category
    if not info:
        info = ['other']
        log_to_ui(f"SESSION_END - No categories detected, defaulting to: {info}")
    
    # Validate and clean categories
    allowed = ['money', 'eligibility', 'district', 'ingredients', 'other']
    info = [i if i in allowed else 'other' for i in info]
    
    # Ensure money is first if present
    if 'money' in info:
        info = ['money'] + [i for i in info if i != 'money']
    
    # FINAL SAFETY CHECK: Ensure we always have at least one category
    if not info:
        info = ['other']
        log_to_ui(f"SESSION_END - Final safety check: forcing 'other' category")
    
    duration = str(len(chat))

    # Log the categorization results clearly
    log_to_ui(f"=== FINAL SESSION CATEGORIZATION ===")
    log_to_ui(f"Session ID: {sid}")
    log_to_ui(f"Total chat messages: {len(chat)}")
    user_messages = [m.get('content', '')[:50] for m in chat if m.get('role') == 'user']
    log_to_ui(f"User messages: {user_messages}")
    log_to_ui(f"Session types map: {session_types_map}")
    log_to_ui(f"Final info categories: {info}")
    log_to_ui(f"Model used throughout: {current_model}")
    log_to_ui(f"=====================================")
    
    # Enhanced logging to app.log file for permanent record
    logging.info(f"SESSION_END_DETAILED [{now_str()}] Session: {sid}, Messages: {len(chat)}, User: {user_messages}, Types: {info}, Model: {current_model}, Duration: {duration}, Total_User_Messages: {len(user_messages)}")
    
    print(f"SESSION_END - Session: {sid}, Messages: {len(chat)}, Types: {info}, Model: {current_model}")

    payload = {
      "tdatetime": now_str(),
      "sessionid": sid,
      "summary":   summary,
      "duration":  duration,
      "types":     json.dumps({
                       "complaints":    [],
                       "info": info
                   }, ensure_ascii=False),
      "chat":      json.dumps(chat, ensure_ascii=False)
    }
    
    # Log what's being sent to API
    print(f"API PAYLOAD - Session: {sid}")
    print(f"  Types being sent: complaints={[]}, info={info}")
    print(f"  Summary: {summary[:100]}...")
    print(f"  Duration: {duration}")
    print(f"  Model used: {current_model}")
    
    logging.info(f"Posting transcript payload: {payload}")

    try:
        res = requests.post(
          "https://urdubot.nettechltd.com/api/transcripts",
          json=payload, timeout=10
        )
        res.raise_for_status()
        status = "uploaded"
        logging.info(f"Transcript uploaded: {res.status_code}")
    except Exception as e:
        logging.error(f"Transcripts POST failed: {e}")
        os.makedirs("Interactions", exist_ok=True)
        path = f"Interactions/session_{sid}_{now_str()}.json"
        with open(path, "w", encoding="utf8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        status = "saved_locally"

    session_type_history.pop(sid, None)
    session_types.pop(sid, None)
    return jsonify({"ended":True, "status":status})

@app.route("/chat", methods=["POST"])
def chat():
    data   = request.get_json() or {}
    sid    = data.get("session_id")
    # Accept both 'message' and 'prompt' for compatibility
    prompt = data.get("message") or data.get("prompt", "")
    session_type = data.get("session_type")
    if sid not in histories:
        return jsonify(response="Please start a session first."), 400

    # Check for matching question first
    matched_answer = find_matching_question(prompt)
    if matched_answer:
        # Store the user message first for proper categorization
        histories[sid].append({"role":"user","content":prompt, "session_type": "info"})
        
        # Categorize the message even for predefined answers
        session_types_map = {len(histories[sid])-1: "info"}
        detected_types = ai_detect_types(histories[sid], session_types_map)
        info = detected_types.get("info", [])
        
        # CRITICAL: Ensure every message is categorized - if no categories found, use 'other'
        if not info:
            info = ['other']
            log_to_ui(f"CHAT (PREDEFINED) - No categories detected, forcing default: {info}")
        
        # Validate and clean categories
        allowed = ['money', 'eligibility', 'district', 'ingredients', 'other']
        info = [i if i in allowed else 'other' for i in info]
        
        # Ensure money is first if present
        if 'money' in info:
            info = ['money'] + [i for i in info if i != 'money']
        
        # FINAL SAFETY CHECK: Ensure we always have at least one category
        if not info:
            info = ['other']
            log_to_ui(f"CHAT (PREDEFINED) - Final safety check: forcing 'other' category")
        
        # Log categorization for predefined answers
        log_to_ui(f"=== CHAT CATEGORIZATION (PREDEFINED) ===")
        log_to_ui(f"Session ID: {sid}")
        log_to_ui(f"User message: '{prompt}'")
        log_to_ui(f"Session type: info")
        log_to_ui(f"Detected types: {detected_types}")
        log_to_ui(f"Final info: {info}")
        log_to_ui(f"Model: predefined_kb")
        log_to_ui(f"=========================================")
        
        # Enhanced logging to app.log file for permanent record
        logging.info(f"PREDEFINED_ANSWER_DETAILED [{now_str()}] Session: {sid}, User: '{prompt[:100]}...', Types: {detected_types}, Final: {info}, Model: predefined_kb, Message_Length: {len(prompt)}")
        
        print(f"CHAT (PREDEFINED) - Session: {sid}, Message: '{prompt[:50]}...', Types: {detected_types}, Final: info={info}")
        
        # Store the assistant response
        histories[sid].append({"role":"assistant","content":matched_answer})
        
        return jsonify(response=matched_answer, kb_excerpt="")

    # Store session type if provided
    if session_type:
        session_types[sid] = session_type
        if sid not in session_type_history:
            session_type_history[sid] = set()
        session_type_history[sid].add(session_type)

    # All sessions are now info sessions
    stype = "info"
    session_types[sid] = stype

    # Store the session_type with each user message
    histories[sid].append({"role":"user","content":prompt, "session_type": stype})

    # Load knowledge base and get relevant chunks using FAISS
    try:
        with open("custom_knowledge.txt", "r", encoding="utf8") as f:
            full_knowledge = f.read()
    except Exception as e:
        full_knowledge = ""

    # Search for relevant knowledge chunks using FAISS
    relevant_chunks = search_knowledge(prompt, top_k=5)  # Increased to 5 for better conflict detection
    
    # Format knowledge with file sources for conflict detection
    if relevant_chunks:
        # Group chunks by file source
        chunks_by_file = {}
        for chunk in relevant_chunks:
            file_name = chunk.get("file_name", "Unknown Source")
            if file_name not in chunks_by_file:
                chunks_by_file[file_name] = []
            chunks_by_file[file_name].append(chunk)
        
        # Format knowledge with clear file sources
        knowledge_parts = []
        for file_name, chunks in chunks_by_file.items():
            file_knowledge = f"From {file_name}:\n"
            for chunk in chunks:
                file_knowledge += f"{chunk['text']}\n"
            knowledge_parts.append(file_knowledge)
        
        relevant_knowledge = "\n\n".join(knowledge_parts)
    else:
        relevant_knowledge = full_knowledge

    # Build session_types_map for ai_detect_types
    session_types_map = {}
    for idx, m in enumerate(histories[sid]):
        if m.get('role') == 'user':
            session_types_map[idx] = m.get('session_type', 'info')
    
    detected_types = ai_detect_types(histories[sid], session_types_map)
    info = detected_types.get("info", [])
    
    # CRITICAL: Ensure every message is categorized - if no categories found, use 'other'
    if not info:
        info = ['other']
        log_to_ui(f"CHAT - No categories detected, forcing default: {info}")
    
    # Validate and clean categories
    allowed = ['money', 'eligibility', 'district', 'ingredients', 'other']
    info = [i if i in allowed else 'other' for i in info]
    
    # Ensure money is first if present
    if 'money' in info:
        info = ['money'] + [i for i in info if i != 'money']
    
    # FINAL SAFETY CHECK: Ensure we always have at least one category
    if not info:
        info = ['other']
        log_to_ui(f"CHAT - Final safety check: forcing 'other' category")
    
    # Log what's being sent to backend
    log_to_ui(f"=== CHAT CATEGORIZATION ===")
    log_to_ui(f"Session ID: {sid}")
    log_to_ui(f"User message: '{prompt}'")
    log_to_ui(f"Session type: {stype}")
    log_to_ui(f"Detected types: {detected_types}")
    log_to_ui(f"Final info: {info}")
    log_to_ui(f"Model being used: {current_model}")
    log_to_ui(f"==========================")
    
    # Enhanced logging to app.log file for permanent record
    logging.info(f"CHAT_SESSION_DETAILED [{now_str()}] Session: {sid}, User: '{prompt[:100]}...', Types: {detected_types}, Final: {info}, Model: {current_model}, Message_Length: {len(prompt)}")
    
    print(f"CHAT - Session: {sid}, Message: '{prompt[:50]}...', Types: {detected_types}, Final: info={info}, Model={current_model}")

    # System prompt for language and direct answers with conflict detection
    system_p = (
        "You are an assistant for the BISP Nashonuma program. "
        "IMPORTANT: Keep answers CONCISE but INFORMATIVE. "
        "Answer the user's question directly using the knowledge base. "
        "Use 2-4 sentences maximum unless more detail is specifically requested. "
        
        "Detect the user's language (English, Urdu, or Roman Urdu) and reply in the same language. "
        "Highlight key information with **bold** formatting. "
        
        "CONFLICT DETECTION: If you find conflicting information in different knowledge base files for the same question, "
        "clearly explain the conflict to the user. Show both pieces of information with their file sources. "
        "Example: 'I found conflicting information in 2 knowledge base files: From file1.pdf: Age requirement is 17-21 years. "
        "From file2.pdf: Age requirement is 13-19 years. Please verify the most current information.' "
        
        "ALWAYS highlight these in **bold**: "
        "- District names (e.g., **Neelam**, **Lasbela**, **Hub**, **Swat**, **Rajanpur**) "
        "- Age requirements (e.g., **13-19 years**) "
        "- Cash amounts (e.g., **Rs. 2000**) "
        "- Document names (e.g., **B-Form**, **CNIC**) "
        "- Program names (e.g., **BISP Nashonuma**) "
        "- Important numbers (e.g., **4 sisters**, **3 months**) "
        "- Key services (e.g., **IFA tablets**) "
        
        "REMEMBER: Be concise but helpful. Use the knowledge base to answer questions accurately."
    )
    
    if current_model == "groq":
        # Use Groq API
        groq_url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {GROQ_API_KEY}"
        }

        msgs = [
            {"role": "system", "content": system_p}
        ]
        if relevant_knowledge:
            msgs.append({"role": "system", "content": relevant_knowledge})
        msgs.append({"role": "user", "content": prompt})

        payload = {
            "model": "llama3-8b-8192",
            "messages": msgs,
            "temperature": 0.4
        }
        
        # Use retry logic for Groq API calls
        resp = make_groq_request_with_retry(groq_url, headers, payload)
        
        if resp is None:
            logging.error("Groq chat API request failed after all retries")
            reply = "Sorry, I'm experiencing technical difficulties. Please try again."
        elif resp.status_code != 200:
            logging.error(f"Groq API error: {resp.status_code} - {resp.text}")
            reply = f"Groq API error: {resp.status_code}"
        else:
            try:
                data = resp.json()
                # Log token usage for Groq
                input_tokens = data.get("usage", {}).get("prompt_tokens", 0)
                output_tokens = data.get("usage", {}).get("completion_tokens", 0)
                log_token_usage("Groq Llama3-8B", input_tokens, output_tokens)
                reply = data["choices"][0]["message"]["content"]
            except Exception as e:
                logging.error(f"Groq response parsing failed: {e}")
                reply = "Sorry, I'm experiencing technical difficulties. Please try again."

    else:
        # Use OpenAI API
        context = [
            {"role":"system","content":system_p},
            {"role":"system","content":relevant_knowledge}
        ] + histories[sid][-20:]
        
        try:
            oa_resp = openai_client.chat.completions.create(
                model="gpt-4",
                     messages=context,
                     temperature=0.4
                   )
            # Log token usage for OpenAI
            input_tokens = oa_resp.usage.prompt_tokens
            output_tokens = oa_resp.usage.completion_tokens
            log_token_usage("OpenAI GPT-4", input_tokens, output_tokens)
            reply = oa_resp.choices[0].message.content
        except Exception as e:
            logging.error(f"OpenAI API error: {e}")
            reply = f"Error: {str(e)}"
    
    # Keep the ** markers for bold formatting - don't remove them
    # reply = reply.replace("**", "", 1).replace("**", " ", 1)

    histories[sid].append({"role":"assistant","content":reply})

    # Get relevant knowledge base excerpt with file sources
    kb_excerpt = ""
    if relevant_chunks:
        # Show top 2 most relevant chunks with file sources
        excerpt_parts = []
        for chunk in relevant_chunks[:2]:
            file_name = chunk.get("file_name", "Unknown Source")
            excerpt_parts.append(f"From {file_name}: {chunk['text'][:200]}...")
        kb_excerpt = "\n\n".join(excerpt_parts)

    return jsonify(response=reply, kb_excerpt=kb_excerpt)

@app.route("/suggestions", methods=["POST"])
def suggestions():
    data = request.get_json() or {}
    user_input = data.get("input", "").lower()
    
    # Get suggestions from predefined questions first
    suggestions = []
    for q in PROGRAM_QA:
        if not user_input or any(word in user_input for word in q.split()):
            suggestions.append(q.capitalize() + "?")
        if len(suggestions) >= 3:
            break
    
    # If we need more suggestions, use Groq (much cheaper than OpenAI)
    if len(suggestions) < 3 and GROQ_API_KEY:
        try:
            with open("custom_knowledge.txt", "r", encoding="utf8") as f:
                knowledge = f.read()
            
            # Optimize the prompt to be shorter and more focused
            system_prompt = (
                "Generate 2-3 SHORT, specific questions based on the knowledge base. "
                "Each question should be 5-10 words maximum. "
                "Focus on common user inquiries about eligibility, money, documents, or locations. "
                "Return ONLY the questions, one per line, no extra text or numbering."
            )
            
            # Truncate knowledge to reduce tokens (limit to first 1000 chars)
            truncated_knowledge = knowledge[:1000] + "..." if len(knowledge) > 1000 else knowledge
            user_content = f"Knowledge: {truncated_knowledge}\nUser: {user_input}"
            
            groq_url = "https://api.groq.com/openai/v1/chat/completions"
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {GROQ_API_KEY}"
            }
            payload = {
                "model": "llama3-8b-8192",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content}
                ],
                "temperature": 0.5,
                "max_tokens": 100  # Limit output tokens
            }
            
            # Use retry logic for Groq API calls
            resp = make_groq_request_with_retry(groq_url, headers, payload)
            
            if resp is None:
                logging.error("Groq suggestion generation failed after all retries")
                text = ""
            elif resp.status_code != 200:
                logging.error(f"Groq suggestion generation failed: {resp.status_code} - {resp.text}")
                text = ""
            else:
                try:
                    data = resp.json()
                    # Log token usage for Groq suggestions
                    input_tokens = data.get("usage", {}).get("prompt_tokens", 0)
                    output_tokens = data.get("usage", {}).get("completion_tokens", 0)
                    log_token_usage("Groq Llama3-8B (suggestions)", input_tokens, output_tokens)
                    text = data["choices"][0]["message"]["content"].strip()
                except Exception as e:
                    logging.error(f"Groq suggestion response parsing failed: {e}")
                    text = ""
            
            if text:
                # Filter out unwanted text and clean suggestions
                lines = text.split('\n')
                ai_suggestions = []
                for line in lines:
                    line = line.strip()
                    if line:
                        # Remove common unwanted prefixes
                        line = line.strip('-* 1234567890.')
                        # Skip lines that are not actual questions
                        skip_phrases = [
                            'here are', 'questions:', 'suggestions:', 'following', 
                            'related questions', 'you might ask', 'consider asking'
                        ]
                        if not any(phrase in line.lower() for phrase in skip_phrases):
                            # Ensure it ends with a question mark
                            if not line.endswith('?'):
                                line += '?'
                            ai_suggestions.append(line)
                
                suggestions.extend(ai_suggestions[:3-len(suggestions)])
        except Exception as e:
            logging.error(f"Suggestion generation failed: {e}")
    
    # Ensure we always return 3 suggestions
    default_suggestions = [
        "What is the age requirement?",
        "How much is the cash transfer?",
        "What documents are needed?"
    ]
    while len(suggestions) < 3:
        suggestions.append(default_suggestions[len(suggestions)])
    
    return jsonify({"suggestions": suggestions[:3]})

@app.route('/logs', methods=['GET'])
def get_logs():
    return jsonify({'logs': RECENT_LOGS[-MAX_LOGS:]})

if __name__ == "__main__":
    # Initialize FAISS
    initialize_faiss()
    
    # Load existing knowledge base into FAISS if available
    try:
        with open("custom_knowledge.txt", "r", encoding="utf8") as f:
            knowledge_text = f.read()
            if knowledge_text.strip():
                update_faiss_index(knowledge_text)
    except Exception as e:
        logging.info("No existing knowledge base found for FAISS initialization")
    
    app.run(host="0.0.0.0", port=5000)